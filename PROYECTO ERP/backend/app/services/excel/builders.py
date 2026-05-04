"""Builders concretos para los exports Excel.

Cada función acepta una sesión SQLAlchemy + filtros y devuelve `bytes` con
el `.xlsx` listo para descargar.

Diseño:
- Una sola query por reporte (no N+1). Joins explícitos contra tablas indexadas.
- Sumas y agregaciones se calculan en SQL cuando es viable, pero para Libro IVA
  Digital también desglosamos por IVA% en Python (queryear todas las líneas).
- Las cabeceras siguen el orden de columnas que el contador espera (RG 4597
  para Libro IVA Digital).

NOTA: Por ahora el sheet "Compras" del Libro IVA queda con la misma lógica
heredada (vacío en la sub-fase OCR), pero el resto de exports de compras se
construye sobre `Factura` tipo `factura_c` (compras vía OCR).
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.alerta import Alerta
from app.models.articulo import Articulo
from app.models.cae import Cae
from app.models.cliente import CondicionIvaEnum, Cliente
from app.models.factura import EstadoComprobanteEnum, Factura, TipoComprobanteEnum
from app.models.factura_item import FacturaItem
from app.models.pago import FacturaPago, MedioPagoEnum
from app.models.proveedor import Proveedor
from app.models.resumen import MovimientoCaja, TipoMovimientoEnum
from app.models.stock import StockSucursal
from app.models.sucursal import Sucursal

# --- Estilos compartidos -----------------------------------------------------

HEADER_FILL = PatternFill(start_color="FFEDEDED", end_color="FFEDEDED", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FF1A1A1A", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

CURRENCY_FMT = '"$"#,##0.00'
DATE_FMT = "yyyy-mm-dd"
INT_FMT = "#,##0"
PERCENT_FMT = "0.00%"

# Tipos de Factura considerados "compra" (vienen del flujo OCR -> proveedor).
COMPRA_TIPOS = (TipoComprobanteEnum.factura_c,)


def _ensure_styles(wb: Workbook) -> None:
    """Registra el NamedStyle currency_ars en el workbook si no existe."""
    if "currency_ars" not in wb.named_styles:
        ns = NamedStyle(name="currency_ars", number_format=CURRENCY_FMT)
        wb.add_named_style(ns)


def _apply_header(ws: Worksheet, headers: list[str]) -> None:
    """Escribe headers en la fila 1 con estilo bold + fill, y freezea."""
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.freeze_panes = "A2"


def _autofit(ws: Worksheet, max_width: int = 40) -> None:
    """Ajusta el ancho de cada columna a la longitud máxima de su contenido."""
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _date_str(value: date | datetime | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value.isoformat()


def _decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _condicion_iva_codigo(c: CondicionIvaEnum | None) -> int:
    """Mapea condición IVA del receptor al código RG 5616 de AFIP."""
    mapping = {
        CondicionIvaEnum.responsable_inscripto: 1,
        CondicionIvaEnum.exento: 4,
        CondicionIvaEnum.consumidor_final: 5,
        CondicionIvaEnum.monotributo: 6,
        CondicionIvaEnum.no_categorizado: 7,
    }
    if c is None:
        return 5
    return mapping.get(c, 5)


def _tipo_comprobante_codigo(tipo: TipoComprobanteEnum) -> int:
    """Mapea tipo de comprobante interno al código numérico AFIP."""
    mapping = {
        TipoComprobanteEnum.factura_a: 1,
        TipoComprobanteEnum.factura_b: 6,
        TipoComprobanteEnum.factura_c: 11,
        TipoComprobanteEnum.ticket: 83,
        TipoComprobanteEnum.nota_credito_a: 3,
        TipoComprobanteEnum.nota_credito_b: 8,
        TipoComprobanteEnum.nota_credito_c: 13,
        TipoComprobanteEnum.remito: 91,
        TipoComprobanteEnum.presupuesto: 0,
    }
    return mapping.get(tipo, 0)


def _legacy_proveedor_id(factura: Factura) -> int | None:
    """Extrae el proveedor_id desde legacy_meta (las compras OCR lo guardan ahí)."""
    if not factura.legacy_meta:
        return None
    raw = factura.legacy_meta.get("proveedor_id")
    if raw is None:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _set_cell(cell, val, fmt: str | None = None, align=None, style: str | None = None) -> None:
    cell.value = val
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if style:
        cell.style = style


# --- Libro IVA Digital -------------------------------------------------------


def build_libro_iva_digital(
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
) -> bytes:
    """Genera el Libro IVA Digital (RG 4597) — sheets Ventas y Compras."""
    wb = Workbook()
    _ensure_styles(wb)

    ventas_ws = wb.active
    ventas_ws.title = "Ventas"
    _build_libro_iva_ventas_sheet(ventas_ws, session, fecha_desde, fecha_hasta)

    compras_ws = wb.create_sheet(title="Compras")
    _build_libro_iva_compras_sheet(compras_ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


VENTAS_HEADERS = [
    "Fecha",
    "Tipo Cbte",
    "Cód. AFIP",
    "Pto. Venta",
    "Número",
    "CUIT",
    "Razón Social",
    "Cond. IVA Recep.",
    "Neto Gravado",
    "IVA 21%",
    "IVA 10.5%",
    "IVA Otras",
    "Otros Tributos",
    "Total",
]


def _build_libro_iva_ventas_sheet(
    ws: Worksheet,
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
) -> None:
    _apply_header(ws, VENTAS_HEADERS)

    desde_dt = datetime.combine(fecha_desde, datetime.min.time())
    hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

    fact_stmt = (
        select(Factura, Cliente)
        .outerjoin(Cliente, Cliente.id == Factura.cliente_id)
        .where(
            Factura.fecha >= desde_dt,
            Factura.fecha <= hasta_dt,
            Factura.estado == EstadoComprobanteEnum.emitida,
            Factura.tipo != TipoComprobanteEnum.factura_c,
        )
        .order_by(Factura.fecha, Factura.id)
    )
    rows = session.execute(fact_stmt).all()

    if not rows:
        return

    factura_ids = [r[0].id for r in rows]

    iva_stmt = (
        select(
            FacturaItem.factura_id.label("factura_id"),
            FacturaItem.iva_porc.label("iva_porc"),
            func.coalesce(func.sum(FacturaItem.subtotal), 0).label("neto"),
            func.coalesce(func.sum(FacturaItem.iva_monto), 0).label("iva"),
        )
        .where(FacturaItem.factura_id.in_(factura_ids))
        .group_by(FacturaItem.factura_id, FacturaItem.iva_porc)
    )
    iva_rows = session.execute(iva_stmt).all()
    iva_by_factura: dict[int, dict[str, Decimal]] = {}
    for r in iva_rows:
        bucket = iva_by_factura.setdefault(
            r.factura_id, {"21": Decimal("0"), "10_5": Decimal("0"), "otras": Decimal("0"), "neto": Decimal("0")}
        )
        bucket["neto"] += _decimal(r.neto)
        porc = _decimal(r.iva_porc)
        if porc == Decimal("21"):
            bucket["21"] += _decimal(r.iva)
        elif porc == Decimal("10.5"):
            bucket["10_5"] += _decimal(r.iva)
        else:
            bucket["otras"] += _decimal(r.iva)

    row_idx = 2
    for factura, cliente in rows:
        breakdown = iva_by_factura.get(
            factura.id,
            {"21": Decimal("0"), "10_5": Decimal("0"), "otras": Decimal("0"), "neto": Decimal("0")},
        )
        cuit = cliente.cuit if cliente else ""
        razon = cliente.razon_social if cliente else "Consumidor Final"
        cond_codigo = _condicion_iva_codigo(cliente.condicion_iva if cliente else None)

        values = [
            factura.fecha.date() if isinstance(factura.fecha, datetime) else factura.fecha,
            factura.tipo.value,
            _tipo_comprobante_codigo(factura.tipo),
            factura.punto_venta,
            factura.numero,
            cuit or "",
            razon,
            cond_codigo,
            float(breakdown["neto"]),
            float(breakdown["21"]),
            float(breakdown["10_5"]),
            float(breakdown["otras"]),
            0.0,
            float(_decimal(factura.total)),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.number_format = DATE_FMT
            elif col_idx in (3, 4, 5, 8):
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx in (9, 10, 11, 12, 13, 14):
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1

    _autofit(ws)


COMPRAS_HEADERS = [
    "Fecha",
    "Tipo Cbte",
    "Cód. AFIP",
    "Pto. Venta",
    "Número",
    "CUIT Proveedor",
    "Razón Social",
    "Neto Gravado",
    "IVA 21%",
    "IVA 10.5%",
    "Otros Tributos",
    "Total",
]


def _build_libro_iva_compras_sheet(ws: Worksheet) -> None:
    """Sheet Compras del Libro IVA Digital — vacío hoy (OCR en construcción)."""
    _apply_header(ws, COMPRAS_HEADERS)
    _autofit(ws)


# --- Ventas detalladas -------------------------------------------------------


VENTAS_EXPORT_HEADERS = [
    "Fecha",
    "Sucursal",
    "Tipo",
    "Pto. Venta",
    "Número",
    "Cliente",
    "CUIT",
    "Cajero",
    "Subtotal",
    "IVA",
    "Total",
    "Estado",
]


def build_ventas_export(
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
    sucursal_id: int | None = None,
) -> bytes:
    """Export detallado de ventas — sheet "Detalle" + sheet "Resumen Diario"."""
    wb = Workbook()
    _ensure_styles(wb)

    detalle_ws = wb.active
    detalle_ws.title = "Detalle"
    detalle_rows = _build_ventas_detalle_sheet(
        detalle_ws, session, fecha_desde, fecha_hasta, sucursal_id
    )

    resumen_ws = wb.create_sheet(title="Resumen Diario")
    _build_ventas_resumen_diario_sheet(resumen_ws, detalle_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_ventas_detalle_sheet(
    ws: Worksheet,
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
    sucursal_id: int | None,
) -> list[dict[str, Any]]:
    from app.models.user import User

    _apply_header(ws, VENTAS_EXPORT_HEADERS)

    desde_dt = datetime.combine(fecha_desde, datetime.min.time())
    hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

    where = [
        Factura.fecha >= desde_dt,
        Factura.fecha <= hasta_dt,
        Factura.estado == EstadoComprobanteEnum.emitida,
        Factura.tipo != TipoComprobanteEnum.factura_c,
    ]
    if sucursal_id:
        where.append(Factura.sucursal_id == sucursal_id)

    stmt = (
        select(Factura, Sucursal, Cliente, User)
        .join(Sucursal, Sucursal.id == Factura.sucursal_id)
        .outerjoin(Cliente, Cliente.id == Factura.cliente_id)
        .join(User, User.id == Factura.cajero_id)
        .where(*where)
        .order_by(Factura.fecha, Factura.id)
    )
    rows = session.execute(stmt).all()

    detalle: list[dict[str, Any]] = []
    row_idx = 2
    for factura, sucursal, cliente, cajero in rows:
        fecha_dia = (
            factura.fecha.date() if isinstance(factura.fecha, datetime) else factura.fecha
        )
        values = [
            fecha_dia,
            f"{sucursal.codigo} - {sucursal.nombre}",
            factura.tipo.value,
            factura.punto_venta,
            factura.numero,
            cliente.razon_social if cliente else "Consumidor Final",
            cliente.cuit if cliente else "",
            cajero.nombre if cajero else "",
            float(_decimal(factura.subtotal)),
            float(_decimal(factura.total_iva)),
            float(_decimal(factura.total)),
            factura.estado.value,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.number_format = DATE_FMT
            elif col_idx in (4, 5):
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx in (9, 10, 11):
                cell.style = "currency_ars"
                cell.alignment = RIGHT

        detalle.append(
            {
                "fecha": fecha_dia,
                "subtotal": _decimal(factura.subtotal),
                "iva": _decimal(factura.total_iva),
                "total": _decimal(factura.total),
            }
        )
        row_idx += 1

    _autofit(ws)
    return detalle


def _build_ventas_resumen_diario_sheet(
    ws: Worksheet, detalle: Iterable[dict[str, Any]]
) -> None:
    headers = ["Fecha", "Comprobantes", "Subtotal", "IVA", "Total"]
    _apply_header(ws, headers)

    by_day: dict[date, dict[str, Decimal | int]] = {}
    for r in detalle:
        bucket = by_day.setdefault(
            r["fecha"],
            {"cantidad": 0, "subtotal": Decimal("0"), "iva": Decimal("0"), "total": Decimal("0")},
        )
        bucket["cantidad"] = int(bucket["cantidad"]) + 1
        bucket["subtotal"] = Decimal(bucket["subtotal"]) + r["subtotal"]
        bucket["iva"] = Decimal(bucket["iva"]) + r["iva"]
        bucket["total"] = Decimal(bucket["total"]) + r["total"]

    row_idx = 2
    for fecha in sorted(by_day.keys()):
        bucket = by_day[fecha]
        values = [
            fecha,
            int(bucket["cantidad"]),
            float(Decimal(bucket["subtotal"])),
            float(Decimal(bucket["iva"])),
            float(Decimal(bucket["total"])),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.number_format = DATE_FMT
            elif col_idx == 2:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx in (3, 4, 5):
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1

    _autofit(ws)


# --- Stock --------------------------------------------------------------------


STOCK_HEADERS = [
    "Código",
    "Descripción",
    "Unidad",
    "Sucursal",
    "Cantidad",
    "Costo unitario",
    "Valor stock",
    "PVP base",
]


def build_stock_export(
    session: Session,
    sucursal_id: int | None = None,
) -> bytes:
    """Export de stock por (artículo × sucursal). Snapshot al momento."""
    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = "Stock"
    _apply_header(ws, STOCK_HEADERS)

    stmt = (
        select(StockSucursal, Articulo, Sucursal)
        .join(Articulo, Articulo.id == StockSucursal.articulo_id)
        .join(Sucursal, Sucursal.id == StockSucursal.sucursal_id)
        .where(Articulo.activo.is_(True))
    )
    if sucursal_id:
        stmt = stmt.where(StockSucursal.sucursal_id == sucursal_id)
    stmt = stmt.order_by(Articulo.codigo, Sucursal.codigo)

    rows = session.execute(stmt).all()

    row_idx = 2
    for stock, articulo, sucursal in rows:
        cantidad = _decimal(stock.cantidad)
        costo = _decimal(articulo.costo)
        valor = cantidad * costo
        values = [
            articulo.codigo,
            articulo.descripcion,
            articulo.unidad_medida.value,
            f"{sucursal.codigo} - {sucursal.nombre}",
            float(cantidad),
            float(costo),
            float(valor),
            float(_decimal(articulo.pvp_base)),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 5:
                cell.number_format = "#,##0.0000"
                cell.alignment = RIGHT
            elif col_idx in (6, 7, 8):
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1

    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Compras (Factura tipo C) ------------------------------------------------


COMPRAS_EXPORT_HEADERS = [
    "Fecha",
    "Comprobante",
    "Sucursal",
    "Proveedor",
    "CUIT",
    "Items",
    "Subtotal",
    "IVA",
    "Total",
    "Estado",
]


def _proveedor_lookup(session: Session) -> dict[int, Proveedor]:
    """Cachea proveedores en memoria. Útil para resolver legacy_meta.proveedor_id."""
    return {p.id: p for p in session.execute(select(Proveedor)).scalars().all()}


def build_compras_export(
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
    proveedor_id: int | None = None,
    sucursal_id: int | None = None,
) -> bytes:
    """Export de compras (Factura tipo C) en rango con filtros."""
    wb = Workbook()
    _ensure_styles(wb)

    detalle_ws = wb.active
    detalle_ws.title = "Detalle compras"
    _apply_header(detalle_ws, COMPRAS_EXPORT_HEADERS)

    desde_dt = datetime.combine(fecha_desde, datetime.min.time())
    hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

    where = [
        Factura.fecha >= desde_dt,
        Factura.fecha <= hasta_dt,
        Factura.estado == EstadoComprobanteEnum.emitida,
        Factura.tipo == TipoComprobanteEnum.factura_c,
    ]
    if sucursal_id:
        where.append(Factura.sucursal_id == sucursal_id)

    stmt = (
        select(Factura, Sucursal)
        .join(Sucursal, Sucursal.id == Factura.sucursal_id)
        .where(*where)
        .order_by(Factura.fecha, Factura.id)
    )
    rows = session.execute(stmt).all()

    proveedores = _proveedor_lookup(session)

    factura_ids = [f.id for f, _ in rows]
    items_count: dict[int, int] = {}
    if factura_ids:
        items_stmt = (
            select(FacturaItem.factura_id, func.count(FacturaItem.id))
            .where(FacturaItem.factura_id.in_(factura_ids))
            .group_by(FacturaItem.factura_id)
        )
        for fid, n in session.execute(items_stmt).all():
            items_count[fid] = int(n)

    row_idx = 2
    by_proveedor: dict[int, dict[str, Any]] = {}
    total_global = Decimal("0")

    for factura, sucursal in rows:
        prov_id = _legacy_proveedor_id(factura)
        if proveedor_id and prov_id != proveedor_id:
            continue
        proveedor = proveedores.get(prov_id) if prov_id else None

        fecha_dia = (
            factura.fecha.date() if isinstance(factura.fecha, datetime) else factura.fecha
        )
        nro = (factura.legacy_meta or {}).get("numero_proveedor") or (
            f"{factura.punto_venta:04d}-{factura.numero:08d}"
        )
        razon = proveedor.razon_social if proveedor else (
            (factura.legacy_meta or {}).get("proveedor_nombre_raw") or "Sin proveedor"
        )
        cuit = proveedor.cuit if proveedor else ""
        total = _decimal(factura.total)
        total_global += total

        values = [
            fecha_dia,
            str(nro),
            f"{sucursal.codigo} - {sucursal.nombre}",
            razon,
            cuit or "",
            items_count.get(factura.id, 0),
            float(_decimal(factura.subtotal)),
            float(_decimal(factura.total_iva)),
            float(total),
            factura.estado.value,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = detalle_ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.number_format = DATE_FMT
            elif col_idx == 6:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx in (7, 8, 9):
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1

        # Resumen por proveedor
        key = prov_id or 0
        bucket = by_proveedor.setdefault(
            key,
            {"razon": razon, "cuit": cuit or "", "cantidad": 0, "total": Decimal("0")},
        )
        bucket["cantidad"] = int(bucket["cantidad"]) + 1
        bucket["total"] = Decimal(bucket["total"]) + total

    _autofit(detalle_ws)

    # Sheet Resumen por proveedor
    resumen_ws = wb.create_sheet(title="Resumen por proveedor")
    _apply_header(
        resumen_ws,
        ["Razón social", "CUIT", "Cantidad facturas", "Total facturado", "% del total"],
    )
    row_idx = 2
    for bucket in sorted(
        by_proveedor.values(), key=lambda b: Decimal(b["total"]), reverse=True
    ):
        total_b = Decimal(bucket["total"])
        porc = float(total_b / total_global) if total_global > 0 else 0.0
        values = [
            bucket["razon"],
            bucket["cuit"],
            int(bucket["cantidad"]),
            float(total_b),
            porc,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = resumen_ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 3:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx == 4:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
            elif col_idx == 5:
                cell.number_format = PERCENT_FMT
                cell.alignment = RIGHT
        row_idx += 1
    _autofit(resumen_ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Cobranzas ---------------------------------------------------------------


COBRANZAS_HEADERS = [
    "Fecha",
    "Cliente",
    "CUIT",
    "Factura ref.",
    "Monto",
    "Medio",
    "Sucursal",
    "Cajero",
    "Descripción",
]


def build_cobranzas_export(
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
    cliente_id: int | None = None,
    sucursal_id: int | None = None,
) -> bytes:
    """Cobranzas (movimientos tipo cobranza) detallados + por medio de pago."""
    from app.models.user import User

    wb = Workbook()
    _ensure_styles(wb)

    ws = wb.active
    ws.title = "Detalle"
    _apply_header(ws, COBRANZAS_HEADERS)

    desde_dt = datetime.combine(fecha_desde, datetime.min.time())
    hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

    where = [
        MovimientoCaja.fecha >= desde_dt,
        MovimientoCaja.fecha <= hasta_dt,
        MovimientoCaja.tipo == TipoMovimientoEnum.cobranza,
    ]
    if sucursal_id:
        where.append(MovimientoCaja.sucursal_id == sucursal_id)
    if cliente_id:
        where.append(MovimientoCaja.cliente_id == cliente_id)

    stmt = (
        select(MovimientoCaja, Cliente, Sucursal, User, Factura)
        .outerjoin(Cliente, Cliente.id == MovimientoCaja.cliente_id)
        .join(Sucursal, Sucursal.id == MovimientoCaja.sucursal_id)
        .outerjoin(User, User.id == MovimientoCaja.user_id)
        .outerjoin(Factura, Factura.id == MovimientoCaja.factura_id)
        .where(*where)
        .order_by(MovimientoCaja.fecha, MovimientoCaja.id)
    )
    rows = session.execute(stmt).all()

    by_medio: dict[str, dict[str, Any]] = {}
    total_global = Decimal("0")

    row_idx = 2
    for mov, cliente, sucursal, user, factura in rows:
        fecha_dia = (
            mov.fecha.date() if isinstance(mov.fecha, datetime) else mov.fecha
        )
        ref = (
            f"{factura.tipo.value} {factura.punto_venta:04d}-{factura.numero:08d}"
            if factura
            else ""
        )
        medio_str = mov.medio.value if mov.medio else "—"
        razon = cliente.razon_social if cliente else "Consumidor Final"
        cuit = cliente.cuit if cliente else ""
        monto = _decimal(mov.monto)
        total_global += monto

        values = [
            fecha_dia,
            razon,
            cuit or "",
            ref,
            float(monto),
            medio_str,
            f"{sucursal.codigo} - {sucursal.nombre}",
            user.nombre if user else "",
            mov.descripcion or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.number_format = DATE_FMT
            elif col_idx == 5:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1

        bucket = by_medio.setdefault(medio_str, {"cantidad": 0, "total": Decimal("0")})
        bucket["cantidad"] = int(bucket["cantidad"]) + 1
        bucket["total"] = Decimal(bucket["total"]) + monto

    _autofit(ws)

    medios_ws = wb.create_sheet(title="Por medio de pago")
    _apply_header(medios_ws, ["Medio", "Cantidad", "Total", "% del total"])
    row_idx = 2
    for medio, bucket in sorted(
        by_medio.items(), key=lambda kv: Decimal(kv[1]["total"]), reverse=True
    ):
        total_b = Decimal(bucket["total"])
        porc = float(total_b / total_global) if total_global > 0 else 0.0
        values = [medio, int(bucket["cantidad"]), float(total_b), porc]
        for col_idx, val in enumerate(values, start=1):
            cell = medios_ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 2:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx == 3:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
            elif col_idx == 4:
                cell.number_format = PERCENT_FMT
                cell.alignment = RIGHT
        row_idx += 1
    _autofit(medios_ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Pagos a proveedores ------------------------------------------------------


PAGOS_HEADERS = [
    "Fecha",
    "Proveedor",
    "CUIT",
    "Factura ref.",
    "Monto",
    "Medio",
    "Sucursal",
    "Usuario",
    "Descripción",
]


def build_pagos_export(
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
    proveedor_id: int | None = None,
    sucursal_id: int | None = None,
) -> bytes:
    """Pagos a proveedores (movimientos tipo pago_proveedor)."""
    from app.models.user import User

    wb = Workbook()
    _ensure_styles(wb)

    ws = wb.active
    ws.title = "Detalle"
    _apply_header(ws, PAGOS_HEADERS)

    desde_dt = datetime.combine(fecha_desde, datetime.min.time())
    hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

    where = [
        MovimientoCaja.fecha >= desde_dt,
        MovimientoCaja.fecha <= hasta_dt,
        MovimientoCaja.tipo == TipoMovimientoEnum.pago_proveedor,
    ]
    if sucursal_id:
        where.append(MovimientoCaja.sucursal_id == sucursal_id)
    if proveedor_id:
        where.append(MovimientoCaja.proveedor_id == proveedor_id)

    stmt = (
        select(MovimientoCaja, Proveedor, Sucursal, User, Factura)
        .outerjoin(Proveedor, Proveedor.id == MovimientoCaja.proveedor_id)
        .join(Sucursal, Sucursal.id == MovimientoCaja.sucursal_id)
        .outerjoin(User, User.id == MovimientoCaja.user_id)
        .outerjoin(Factura, Factura.id == MovimientoCaja.factura_id)
        .where(*where)
        .order_by(MovimientoCaja.fecha, MovimientoCaja.id)
    )
    rows = session.execute(stmt).all()

    by_proveedor: dict[int, dict[str, Any]] = {}

    row_idx = 2
    for mov, proveedor, sucursal, user, factura in rows:
        fecha_dia = (
            mov.fecha.date() if isinstance(mov.fecha, datetime) else mov.fecha
        )
        ref = (
            f"{factura.tipo.value} {factura.punto_venta:04d}-{factura.numero:08d}"
            if factura
            else ""
        )
        medio_str = mov.medio.value if mov.medio else "—"
        razon = proveedor.razon_social if proveedor else "Sin proveedor"
        cuit = proveedor.cuit if proveedor else ""
        monto = _decimal(mov.monto)

        values = [
            fecha_dia,
            razon,
            cuit or "",
            ref,
            float(monto),
            medio_str,
            f"{sucursal.codigo} - {sucursal.nombre}",
            user.nombre if user else "",
            mov.descripcion or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.number_format = DATE_FMT
            elif col_idx == 5:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1

        prov_id = proveedor.id if proveedor else 0
        bucket = by_proveedor.setdefault(
            prov_id,
            {"razon": razon, "cuit": cuit or "", "cantidad": 0, "total": Decimal("0")},
        )
        bucket["cantidad"] = int(bucket["cantidad"]) + 1
        bucket["total"] = Decimal(bucket["total"]) + monto

    _autofit(ws)

    prov_ws = wb.create_sheet(title="Por proveedor")
    _apply_header(prov_ws, ["Razón social", "CUIT", "Cantidad pagos", "Total"])
    row_idx = 2
    for bucket in sorted(
        by_proveedor.values(), key=lambda b: Decimal(b["total"]), reverse=True
    ):
        values = [
            bucket["razon"],
            bucket["cuit"],
            int(bucket["cantidad"]),
            float(Decimal(bucket["total"])),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = prov_ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 3:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx == 4:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1
    _autofit(prov_ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Cuenta corriente cliente ------------------------------------------------


def _cliente_movs_running_balance(
    session: Session,
    cliente_id: int,
    fecha_desde: date | None,
    fecha_hasta: date | None,
) -> tuple[list[dict[str, Any]], Decimal]:
    """Devuelve los movimientos de un cliente con saldo running. Saldo positivo = el cliente debe."""
    where = [MovimientoCaja.cliente_id == cliente_id]
    if fecha_desde:
        where.append(MovimientoCaja.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        where.append(MovimientoCaja.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))

    stmt = (
        select(MovimientoCaja, Factura)
        .outerjoin(Factura, Factura.id == MovimientoCaja.factura_id)
        .where(*where)
        .order_by(MovimientoCaja.fecha, MovimientoCaja.id)
    )
    rows = session.execute(stmt).all()
    salida: list[dict[str, Any]] = []
    saldo = Decimal("0")
    for mov, factura in rows:
        debe = Decimal("0")
        haber = Decimal("0")
        descripcion = mov.descripcion or ""
        # Convención: venta a cuenta corriente = debe (cliente debe). cobranza = haber.
        if mov.tipo == TipoMovimientoEnum.venta:
            debe = _decimal(mov.monto)
            if not descripcion and factura is not None:
                descripcion = (
                    f"Venta {factura.tipo.value} "
                    f"{factura.punto_venta:04d}-{factura.numero:08d}"
                )
        elif mov.tipo == TipoMovimientoEnum.cobranza:
            haber = _decimal(mov.monto)
            if not descripcion:
                descripcion = "Cobranza"
        elif mov.tipo == TipoMovimientoEnum.devolucion:
            haber = _decimal(mov.monto)
            if not descripcion:
                descripcion = "Devolución"
        elif mov.tipo == TipoMovimientoEnum.ajuste:
            # ajuste → si monto > 0 lo aplicamos como debe; <0 como haber.
            m = _decimal(mov.monto)
            if m >= 0:
                debe = m
            else:
                haber = -m
            if not descripcion:
                descripcion = "Ajuste"
        else:
            # Otros tipos no afectan cta cte cliente.
            continue

        saldo += debe - haber
        salida.append(
            {
                "fecha": mov.fecha.date() if isinstance(mov.fecha, datetime) else mov.fecha,
                "descripcion": descripcion,
                "debe": float(debe),
                "haber": float(haber),
                "saldo": float(saldo),
            }
        )
    return salida, saldo


def build_cta_cte_cliente_export(
    session: Session,
    cliente_id: int,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
) -> bytes:
    """Histórico de cuenta corriente de un cliente, con saldo running."""
    cliente = session.get(Cliente, cliente_id)
    if cliente is None:
        raise ValueError(f"Cliente {cliente_id} no encontrado")

    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = "Histórico"

    # Header con datos del cliente (filas 1-3)
    ws.cell(row=1, column=1, value="Cliente").font = HEADER_FONT
    ws.cell(row=1, column=2, value=f"{cliente.codigo} · {cliente.razon_social}")
    ws.cell(row=2, column=1, value="CUIT").font = HEADER_FONT
    ws.cell(row=2, column=2, value=cliente.cuit or "—")
    ws.cell(row=3, column=1, value="Saldo cuenta").font = HEADER_FONT
    ws.cell(row=3, column=2, value=float(_decimal(cliente.saldo))).style = "currency_ars"

    # Headers de la tabla (fila 5)
    headers = ["Fecha", "Descripción", "Debe", "Haber", "Saldo"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.freeze_panes = "A6"

    movs, saldo_final = _cliente_movs_running_balance(
        session, cliente_id, fecha_desde, fecha_hasta
    )

    row_idx = 6
    for m in movs:
        ws.cell(row=row_idx, column=1, value=m["fecha"]).number_format = DATE_FMT
        ws.cell(row=row_idx, column=2, value=m["descripcion"])
        c_debe = ws.cell(row=row_idx, column=3, value=m["debe"])
        c_debe.style = "currency_ars"
        c_debe.alignment = RIGHT
        c_haber = ws.cell(row=row_idx, column=4, value=m["haber"])
        c_haber.style = "currency_ars"
        c_haber.alignment = RIGHT
        c_saldo = ws.cell(row=row_idx, column=5, value=m["saldo"])
        c_saldo.style = "currency_ars"
        c_saldo.alignment = RIGHT
        row_idx += 1

    # Total
    if movs:
        ws.cell(row=row_idx + 1, column=2, value="Saldo final del período").font = HEADER_FONT
        c_total = ws.cell(row=row_idx + 1, column=5, value=float(saldo_final))
        c_total.style = "currency_ars"
        c_total.font = HEADER_FONT
        c_total.alignment = RIGHT

    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cta_cte_proveedor_export(
    session: Session,
    proveedor_id: int,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
) -> bytes:
    """Histórico de cuenta corriente de un proveedor, con saldo running."""
    proveedor = session.get(Proveedor, proveedor_id)
    if proveedor is None:
        raise ValueError(f"Proveedor {proveedor_id} no encontrado")

    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = "Histórico"

    ws.cell(row=1, column=1, value="Proveedor").font = HEADER_FONT
    ws.cell(row=1, column=2, value=f"{proveedor.codigo} · {proveedor.razon_social}")
    ws.cell(row=2, column=1, value="CUIT").font = HEADER_FONT
    ws.cell(row=2, column=2, value=proveedor.cuit or "—")

    headers = ["Fecha", "Descripción", "Debe (compra)", "Haber (pago)", "Saldo"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.freeze_panes = "A5"

    where_movs = [MovimientoCaja.proveedor_id == proveedor_id]
    where_facts: list[Any] = []
    if fecha_desde:
        where_movs.append(MovimientoCaja.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
        where_facts.append(Factura.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        where_movs.append(MovimientoCaja.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))
        where_facts.append(Factura.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))

    movs = session.execute(
        select(MovimientoCaja).where(*where_movs).order_by(MovimientoCaja.fecha, MovimientoCaja.id)
    ).scalars().all()

    # Compras: facturas C cuyo legacy_meta.proveedor_id == proveedor_id
    facts_q = select(Factura).where(
        Factura.tipo == TipoComprobanteEnum.factura_c,
        Factura.estado == EstadoComprobanteEnum.emitida,
        *where_facts,
    )
    facts = [
        f for f in session.execute(facts_q).scalars().all()
        if _legacy_proveedor_id(f) == proveedor_id
    ]

    # Mezclamos eventos por fecha para saldo running
    eventos: list[tuple[datetime, Decimal, Decimal, str]] = []
    for f in facts:
        eventos.append(
            (
                f.fecha,
                _decimal(f.total),  # debe (debemos al proveedor)
                Decimal("0"),
                f"Compra {f.tipo.value} "
                + (f.legacy_meta or {}).get(
                    "numero_proveedor",
                    f"{f.punto_venta:04d}-{f.numero:08d}",
                ),
            )
        )
    for mov in movs:
        if mov.tipo == TipoMovimientoEnum.pago_proveedor:
            eventos.append((mov.fecha, Decimal("0"), _decimal(mov.monto), mov.descripcion or "Pago"))
        elif mov.tipo == TipoMovimientoEnum.ajuste:
            m = _decimal(mov.monto)
            if m >= 0:
                eventos.append((mov.fecha, m, Decimal("0"), mov.descripcion or "Ajuste"))
            else:
                eventos.append((mov.fecha, Decimal("0"), -m, mov.descripcion or "Ajuste"))

    eventos.sort(key=lambda e: e[0])

    saldo = Decimal("0")
    row_idx = 5
    for fecha_e, debe, haber, descripcion in eventos:
        saldo += debe - haber
        ws.cell(row=row_idx, column=1, value=fecha_e.date() if isinstance(fecha_e, datetime) else fecha_e).number_format = DATE_FMT
        ws.cell(row=row_idx, column=2, value=descripcion)
        c1 = ws.cell(row=row_idx, column=3, value=float(debe))
        c1.style = "currency_ars"
        c1.alignment = RIGHT
        c2 = ws.cell(row=row_idx, column=4, value=float(haber))
        c2.style = "currency_ars"
        c2.alignment = RIGHT
        c3 = ws.cell(row=row_idx, column=5, value=float(saldo))
        c3.style = "currency_ars"
        c3.alignment = RIGHT
        row_idx += 1

    if eventos:
        ws.cell(row=row_idx + 1, column=2, value="Saldo final").font = HEADER_FONT
        c_total = ws.cell(row=row_idx + 1, column=5, value=float(saldo))
        c_total.style = "currency_ars"
        c_total.font = HEADER_FONT
        c_total.alignment = RIGHT

    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Resúmenes globales ------------------------------------------------------


def build_resumen_clientes(session: Session) -> bytes:
    """Listado completo de clientes con saldo y última operación."""
    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = "Clientes"
    headers = [
        "Código",
        "Razón social",
        "CUIT",
        "Cond. IVA",
        "Cta. Cte.",
        "Límite",
        "Saldo actual",
        "Última operación",
        "Email",
        "Teléfono",
        "Estado",
    ]
    _apply_header(ws, headers)

    last_mov_stmt = (
        select(
            MovimientoCaja.cliente_id,
            func.max(MovimientoCaja.fecha).label("ultima"),
        )
        .where(MovimientoCaja.cliente_id.is_not(None))
        .group_by(MovimientoCaja.cliente_id)
    )
    ultimas: dict[int, datetime] = {
        cid: ultima for cid, ultima in session.execute(last_mov_stmt).all()
    }

    clientes = session.execute(
        select(Cliente).where(Cliente.deleted_at.is_(None)).order_by(Cliente.razon_social)
    ).scalars().all()

    row_idx = 2
    for c in clientes:
        ultima = ultimas.get(c.id)
        ultima_dia = (
            ultima.date() if isinstance(ultima, datetime) else ultima
        ) if ultima else None
        values = [
            c.codigo,
            c.razon_social,
            c.cuit or "",
            c.condicion_iva.value,
            "Sí" if c.cuenta_corriente else "No",
            float(_decimal(c.limite_cuenta_corriente)),
            float(_decimal(c.saldo)),
            ultima_dia,
            c.email or "",
            c.telefono or "",
            "Activo" if c.activo else "Inactivo",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in (6, 7):
                cell.style = "currency_ars"
                cell.alignment = RIGHT
            elif col_idx == 8:
                cell.number_format = DATE_FMT
        row_idx += 1
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_resumen_proveedores(session: Session) -> bytes:
    """Listado de proveedores con compras totales y último ingreso."""
    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = "Proveedores"
    headers = [
        "Código",
        "Razón social",
        "CUIT",
        "Compras totales",
        "Cantidad facturas",
        "Última compra",
        "Email",
        "Teléfono",
        "Estado",
    ]
    _apply_header(ws, headers)

    proveedores = session.execute(
        select(Proveedor).where(Proveedor.deleted_at.is_(None)).order_by(Proveedor.razon_social)
    ).scalars().all()

    # Sumamos compras (Factura tipo C) por proveedor desde legacy_meta
    facts = session.execute(
        select(Factura).where(
            Factura.tipo == TipoComprobanteEnum.factura_c,
            Factura.estado == EstadoComprobanteEnum.emitida,
        )
    ).scalars().all()
    by_prov: dict[int, dict[str, Any]] = {}
    for f in facts:
        pid = _legacy_proveedor_id(f)
        if pid is None:
            continue
        bucket = by_prov.setdefault(
            pid,
            {"total": Decimal("0"), "cantidad": 0, "ultima": None},
        )
        bucket["total"] = Decimal(bucket["total"]) + _decimal(f.total)
        bucket["cantidad"] = int(bucket["cantidad"]) + 1
        if bucket["ultima"] is None or f.fecha > bucket["ultima"]:
            bucket["ultima"] = f.fecha

    row_idx = 2
    for p in proveedores:
        bucket = by_prov.get(p.id, {"total": Decimal("0"), "cantidad": 0, "ultima": None})
        ultima = bucket["ultima"]
        ultima_dia = (
            ultima.date() if isinstance(ultima, datetime) else ultima
        ) if ultima else None
        values = [
            p.codigo,
            p.razon_social,
            p.cuit or "",
            float(Decimal(bucket["total"])),
            int(bucket["cantidad"]),
            ultima_dia,
            p.email or "",
            p.telefono or "",
            "Activo" if p.activo else "Inactivo",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 4:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
            elif col_idx == 5:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx == 6:
                cell.number_format = DATE_FMT
        row_idx += 1
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Stock valorizado --------------------------------------------------------


def build_stock_valorizado(
    session: Session,
    sucursal_id: int | None = None,
) -> bytes:
    """Stock con costo, valor total y % del valor total."""
    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = "Stock valorizado"
    headers = [
        "Código",
        "Descripción",
        "Familia",
        "Rubro",
        "Marca",
        "Sucursal",
        "Cantidad",
        "Costo unitario",
        "Valor stock",
        "% del valor",
        "PVP base",
    ]
    _apply_header(ws, headers)

    from app.models.categorias import Familia, Marca, Rubro

    stmt = (
        select(StockSucursal, Articulo, Sucursal, Familia, Rubro, Marca)
        .join(Articulo, Articulo.id == StockSucursal.articulo_id)
        .join(Sucursal, Sucursal.id == StockSucursal.sucursal_id)
        .outerjoin(Familia, Familia.id == Articulo.familia_id)
        .outerjoin(Rubro, Rubro.id == Articulo.rubro_id)
        .outerjoin(Marca, Marca.id == Articulo.marca_id)
        .where(Articulo.activo.is_(True))
    )
    if sucursal_id:
        stmt = stmt.where(StockSucursal.sucursal_id == sucursal_id)
    stmt = stmt.order_by(Articulo.codigo, Sucursal.codigo)

    rows = session.execute(stmt).all()

    # Pre-calc valor total para %
    total_valor = Decimal("0")
    enriched: list[dict[str, Any]] = []
    for stock, articulo, sucursal, familia, rubro, marca in rows:
        cantidad = _decimal(stock.cantidad)
        costo = _decimal(articulo.costo)
        valor = cantidad * costo
        total_valor += valor
        enriched.append(
            {
                "stock": stock,
                "articulo": articulo,
                "sucursal": sucursal,
                "familia": familia,
                "rubro": rubro,
                "marca": marca,
                "valor": valor,
                "cantidad": cantidad,
                "costo": costo,
            }
        )

    row_idx = 2
    for r in enriched:
        articulo = r["articulo"]
        sucursal = r["sucursal"]
        porc = (
            float(r["valor"] / total_valor) if total_valor > 0 else 0.0
        )
        values = [
            articulo.codigo,
            articulo.descripcion,
            r["familia"].nombre if r["familia"] else "",
            r["rubro"].nombre if r["rubro"] else "",
            r["marca"].nombre if r["marca"] else "",
            f"{sucursal.codigo} - {sucursal.nombre}",
            float(r["cantidad"]),
            float(r["costo"]),
            float(r["valor"]),
            porc,
            float(_decimal(articulo.pvp_base)),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 7:
                cell.number_format = "#,##0.0000"
                cell.alignment = RIGHT
            elif col_idx in (8, 9, 11):
                cell.style = "currency_ars"
                cell.alignment = RIGHT
            elif col_idx == 10:
                cell.number_format = PERCENT_FMT
                cell.alignment = RIGHT
        row_idx += 1

    # Total
    if enriched:
        ws.cell(row=row_idx + 1, column=8, value="TOTAL").font = HEADER_FONT
        c_total = ws.cell(row=row_idx + 1, column=9, value=float(total_valor))
        c_total.style = "currency_ars"
        c_total.font = HEADER_FONT
        c_total.alignment = RIGHT

    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Ventas detallado (multi-sheet con groupby) ------------------------------


def build_ventas_detallado(
    session: Session,
    fecha_desde: date,
    fecha_hasta: date,
    sucursal_id: int | None = None,
) -> bytes:
    """Ventas detallado con varios sheets de agrupación."""
    from app.models.user import User

    wb = Workbook()
    _ensure_styles(wb)

    desde_dt = datetime.combine(fecha_desde, datetime.min.time())
    hasta_dt = datetime.combine(fecha_hasta, datetime.max.time())

    where = [
        Factura.fecha >= desde_dt,
        Factura.fecha <= hasta_dt,
        Factura.estado == EstadoComprobanteEnum.emitida,
        Factura.tipo != TipoComprobanteEnum.factura_c,
    ]
    if sucursal_id:
        where.append(Factura.sucursal_id == sucursal_id)

    # Sheet 1: Detalle (reutilizamos el builder existente)
    detalle_ws = wb.active
    detalle_ws.title = "Detalle"
    detalle_rows = _build_ventas_detalle_sheet(
        detalle_ws, session, fecha_desde, fecha_hasta, sucursal_id
    )

    # Sheet 2: Por sucursal
    suc_ws = wb.create_sheet(title="Por sucursal")
    _apply_header(suc_ws, ["Sucursal", "Cantidad", "Total"])
    suc_stmt = (
        select(
            Sucursal.codigo,
            Sucursal.nombre,
            func.count(Factura.id).label("cantidad"),
            func.coalesce(func.sum(Factura.total), 0).label("total"),
        )
        .select_from(Factura)
        .join(Sucursal, Sucursal.id == Factura.sucursal_id)
        .where(*where)
        .group_by(Sucursal.codigo, Sucursal.nombre)
        .order_by(func.sum(Factura.total).desc())
    )
    row_idx = 2
    for r in session.execute(suc_stmt).all():
        suc_ws.cell(row=row_idx, column=1, value=f"{r.codigo} - {r.nombre}")
        c1 = suc_ws.cell(row=row_idx, column=2, value=int(r.cantidad or 0))
        c1.number_format = INT_FMT
        c1.alignment = RIGHT
        c2 = suc_ws.cell(row=row_idx, column=3, value=float(_decimal(r.total)))
        c2.style = "currency_ars"
        c2.alignment = RIGHT
        row_idx += 1
    _autofit(suc_ws)

    # Sheet 3: Por familia/rubro
    from app.models.categorias import Familia, Rubro

    fam_ws = wb.create_sheet(title="Por familia-rubro")
    _apply_header(fam_ws, ["Familia", "Rubro", "Unidades", "Total"])
    fam_stmt = (
        select(
            Familia.nombre.label("familia"),
            Rubro.nombre.label("rubro"),
            func.coalesce(func.sum(FacturaItem.cantidad), 0).label("unidades"),
            func.coalesce(func.sum(FacturaItem.total), 0).label("total"),
        )
        .select_from(FacturaItem)
        .join(Factura, Factura.id == FacturaItem.factura_id)
        .join(Articulo, Articulo.id == FacturaItem.articulo_id)
        .outerjoin(Familia, Familia.id == Articulo.familia_id)
        .outerjoin(Rubro, Rubro.id == Articulo.rubro_id)
        .where(*where)
        .group_by(Familia.nombre, Rubro.nombre)
        .order_by(func.sum(FacturaItem.total).desc())
    )
    row_idx = 2
    for r in session.execute(fam_stmt).all():
        fam_ws.cell(row=row_idx, column=1, value=r.familia or "Sin familia")
        fam_ws.cell(row=row_idx, column=2, value=r.rubro or "Sin rubro")
        c1 = fam_ws.cell(row=row_idx, column=3, value=float(_decimal(r.unidades)))
        c1.number_format = "#,##0.00"
        c1.alignment = RIGHT
        c2 = fam_ws.cell(row=row_idx, column=4, value=float(_decimal(r.total)))
        c2.style = "currency_ars"
        c2.alignment = RIGHT
        row_idx += 1
    _autofit(fam_ws)

    # Sheet 4: Por cajero
    cajero_ws = wb.create_sheet(title="Por cajero")
    _apply_header(cajero_ws, ["Cajero", "Cantidad", "Total"])
    cajero_stmt = (
        select(
            User.nombre.label("nombre"),
            func.count(Factura.id).label("cantidad"),
            func.coalesce(func.sum(Factura.total), 0).label("total"),
        )
        .select_from(Factura)
        .join(User, User.id == Factura.cajero_id)
        .where(*where)
        .group_by(User.nombre)
        .order_by(func.sum(Factura.total).desc())
    )
    row_idx = 2
    for r in session.execute(cajero_stmt).all():
        cajero_ws.cell(row=row_idx, column=1, value=r.nombre or "—")
        c1 = cajero_ws.cell(row=row_idx, column=2, value=int(r.cantidad or 0))
        c1.number_format = INT_FMT
        c1.alignment = RIGHT
        c2 = cajero_ws.cell(row=row_idx, column=3, value=float(_decimal(r.total)))
        c2.style = "currency_ars"
        c2.alignment = RIGHT
        row_idx += 1
    _autofit(cajero_ws)

    # Sheet 5: Top 50 productos
    top_ws = wb.create_sheet(title="Top 50 productos")
    _apply_header(top_ws, ["Código", "Descripción", "Unidades vendidas", "Total facturado"])
    top_stmt = (
        select(
            Articulo.codigo,
            Articulo.descripcion,
            func.coalesce(func.sum(FacturaItem.cantidad), 0).label("unidades"),
            func.coalesce(func.sum(FacturaItem.total), 0).label("total"),
        )
        .select_from(FacturaItem)
        .join(Factura, Factura.id == FacturaItem.factura_id)
        .join(Articulo, Articulo.id == FacturaItem.articulo_id)
        .where(*where)
        .group_by(Articulo.codigo, Articulo.descripcion)
        .order_by(func.sum(FacturaItem.cantidad).desc())
        .limit(50)
    )
    row_idx = 2
    for r in session.execute(top_stmt).all():
        top_ws.cell(row=row_idx, column=1, value=r.codigo)
        top_ws.cell(row=row_idx, column=2, value=r.descripcion)
        c1 = top_ws.cell(row=row_idx, column=3, value=float(_decimal(r.unidades)))
        c1.number_format = "#,##0.00"
        c1.alignment = RIGHT
        c2 = top_ws.cell(row=row_idx, column=4, value=float(_decimal(r.total)))
        c2.style = "currency_ars"
        c2.alignment = RIGHT
        row_idx += 1
    _autofit(top_ws)

    # Avoid unused-var warning
    _ = detalle_rows

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Genérico: export Excel desde lista de dicts (para Consultas) ------------


def build_generic_export(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    column_formats: dict[int, str] | None = None,
) -> bytes:
    """Export Excel genérico para la pantalla Consultas. Recibe filas ya proyectadas.

    column_formats: mapping 1-based de columna → formato. Use 'currency', 'date', 'int', 'percent'.
    """
    wb = Workbook()
    _ensure_styles(wb)
    ws = wb.active
    ws.title = title[:31]
    _apply_header(ws, headers)

    fmts = column_formats or {}

    row_idx = 2
    for row in rows:
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            kind = fmts.get(col_idx)
            if kind == "currency":
                cell.style = "currency_ars"
                cell.alignment = RIGHT
            elif kind == "date":
                cell.number_format = DATE_FMT
            elif kind == "int":
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif kind == "percent":
                cell.number_format = PERCENT_FMT
                cell.alignment = RIGHT
        row_idx += 1
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Lista de reposición (stock inteligente — opción C) -----------------------


LISTA_REPOSICION_HEADERS = [
    "Código",
    "Descripción",
    "Sucursal",
    "Stock actual",
    "Mínimo",
    "Reorden",
    "Sugerido pedir",
    "Costo unitario",
    "Total",
    "Urgencia",
]


def build_lista_reposicion(
    session: Session,
    sucursal_id: int | None = None,
) -> bytes:
    """Lista de reposición agrupada por proveedor — para mandar al proveedor.

    Una sheet por proveedor, más una sheet "Resumen" con el total por proveedor.
    Si un proveedor no tiene nombre se etiqueta como "Sin proveedor".
    """
    from app.services.analytics.sugerencias_reposicion import sugerir_reposicion

    data = sugerir_reposicion(session, sucursal_id=sucursal_id)
    grupos = data["por_proveedor"]

    wb = Workbook()
    _ensure_styles(wb)

    # Sheet "Resumen"
    resumen_ws = wb.active
    resumen_ws.title = "Resumen"
    _apply_header(
        resumen_ws,
        ["Proveedor", "CUIT", "Email", "Items a pedir", "Total estimado"],
    )
    row_idx = 2
    for grupo in grupos:
        prov = grupo["proveedor"] or {}
        values = [
            prov.get("razon_social", "Sin proveedor"),
            prov.get("cuit", "") or "",
            prov.get("email", "") or "",
            grupo["total_items"],
            float(_decimal(grupo["total_estimado"])),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = resumen_ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 4:
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            elif col_idx == 5:
                cell.style = "currency_ars"
                cell.alignment = RIGHT
        row_idx += 1
    _autofit(resumen_ws)

    # Una sheet por proveedor
    used_titles: set[str] = {"Resumen"}
    for grupo in grupos:
        prov = grupo["proveedor"] or {}
        raw = prov.get("razon_social") or "Sin proveedor"
        # openpyxl: title <= 31 chars y sin ciertos caracteres
        title = "".join(c for c in raw if c not in "\\/?*[]:")[:31] or "Proveedor"
        # Asegurar único
        base = title
        n = 1
        while title in used_titles:
            n += 1
            sufijo = f" ({n})"
            title = (base[: 31 - len(sufijo)] + sufijo)
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        _apply_header(ws, LISTA_REPOSICION_HEADERS)
        row_idx = 2
        for item in grupo["items"]:
            art = item["articulo"]
            suc = item["sucursal"]
            values = [
                art["codigo"],
                art["descripcion"],
                f"{suc.get('codigo', '')} - {suc.get('nombre', '')}".strip(" -"),
                float(_decimal(item["cantidad_actual"])),
                float(_decimal(item["stock_minimo"])),
                float(_decimal(item["punto_reorden"])),
                float(_decimal(item["cantidad_a_pedir"])),
                float(_decimal(item["costo_unitario"])),
                float(_decimal(item["total_linea"])),
                item["urgencia"],
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx in (4, 5, 6, 7):
                    cell.number_format = "#,##0"
                    cell.alignment = RIGHT
                elif col_idx in (8, 9):
                    cell.style = "currency_ars"
                    cell.alignment = RIGHT
            row_idx += 1
        # Línea de total
        if grupo["items"]:
            total_cell = ws.cell(
                row=row_idx, column=8, value="TOTAL"
            )
            total_cell.font = Font(bold=True)
            total_cell.alignment = RIGHT
            tot = ws.cell(
                row=row_idx,
                column=9,
                value=float(_decimal(grupo["total_estimado"])),
            )
            tot.font = Font(bold=True)
            tot.style = "currency_ars"
            tot.alignment = RIGHT
        _autofit(ws)

    # Si no hay grupos, agregamos sheet vacío informativo
    if not grupos:
        ws = wb.create_sheet(title="Sin reposición")
        ws.cell(row=1, column=1, value="No hay artículos en reposición.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Helper público para que el endpoint construya nombres de archivo coherentes
def make_filename(prefix: str, fecha_desde: date | None = None, fecha_hasta: date | None = None) -> str:
    today = date.today().isoformat()
    if fecha_desde and fecha_hasta:
        return f"{prefix}_{fecha_desde.isoformat()}_{fecha_hasta.isoformat()}.xlsx"
    return f"{prefix}_{today}.xlsx"


__all__ = [
    "build_compras_export",
    "build_cobranzas_export",
    "build_cta_cte_cliente_export",
    "build_cta_cte_proveedor_export",
    "build_generic_export",
    "build_libro_iva_digital",
    "build_lista_reposicion",
    "build_pagos_export",
    "build_resumen_clientes",
    "build_resumen_proveedores",
    "build_stock_export",
    "build_stock_valorizado",
    "build_ventas_detallado",
    "build_ventas_export",
    "make_filename",
]


# Quick sanity import check (no-op at runtime)
_ = timedelta
_ = Cae
_ = Alerta
_ = MedioPagoEnum
