"""Tests del blueprint Consultas (F3) — endpoints genéricos paginados + xlsx."""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.models.articulo import Articulo, UnidadMedidaEnum
from app.models.cliente import Cliente, CondicionIvaEnum
from app.models.factura import EstadoComprobanteEnum, Factura, TipoComprobanteEnum
from app.models.factura_item import FacturaItem
from app.models.pago import FacturaPago, MedioPagoEnum
from app.models.proveedor import Proveedor
from app.models.resumen import MovimientoCaja, TipoMovimientoEnum
from app.models.stock import StockSucursal
from app.models.sucursal import Sucursal

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def contador_user(db):
    from app.models.user import RolEnum, User
    from app.services.auth_service import hash_password

    user = User(
        email="contador-cons@test.example",
        password_hash=hash_password("contador123"),
        nombre="Contador Test",
        rol=RolEnum.contador,
        activo=True,
    )
    db.session.add(user)
    db.session.commit()
    return user


@pytest.fixture
def contador_token(client, contador_user):
    r = client.post(
        "/api/v1/auth/login",
        json={"email": "contador-cons@test.example", "password": "contador123"},
    )
    assert r.status_code == 200, r.get_json()
    return r.get_json()["access_token"]


@pytest.fixture
def sucursal_uno(db):
    s = Sucursal(codigo="S1", nombre="Central", activa=True)
    db.session.add(s)
    db.session.commit()
    return s


@pytest.fixture
def cliente_uno(db):
    c = Cliente(
        codigo="CC01",
        razon_social="Cliente Consulta",
        cuit="20111111111",
        condicion_iva=CondicionIvaEnum.responsable_inscripto,
        cuenta_corriente=True,
        saldo=Decimal("12500.50"),
    )
    db.session.add(c)
    db.session.commit()
    return c


@pytest.fixture
def proveedor_uno(db):
    p = Proveedor(codigo="P01", razon_social="Prov SA", cuit="30222222229", activo=True)
    db.session.add(p)
    db.session.commit()
    return p


@pytest.fixture
def articulo_uno(db, sucursal_uno):
    a = Articulo(
        codigo="ART001",
        descripcion="Articulo Test",
        unidad_medida=UnidadMedidaEnum.unidad,
        costo=Decimal("100"),
        pvp_base=Decimal("200"),
        iva_porc=Decimal("21"),
        activo=True,
        controla_stock=True,
    )
    db.session.add(a)
    db.session.flush()
    db.session.add(StockSucursal(articulo_id=a.id, sucursal_id=sucursal_uno.id, cantidad=Decimal("2")))
    db.session.commit()
    return a


@pytest.fixture
def venta_uno(db, admin_user, sucursal_uno, cliente_uno, articulo_uno):
    fecha = datetime.now(timezone.utc).replace(microsecond=0) - timedelta(days=2)
    f = Factura(
        sucursal_id=sucursal_uno.id,
        punto_venta=1,
        tipo=TipoComprobanteEnum.factura_b,
        numero=1,
        fecha=fecha,
        cliente_id=cliente_uno.id,
        cajero_id=admin_user.id,
        estado=EstadoComprobanteEnum.emitida,
        subtotal=Decimal("1000"),
        total_iva=Decimal("210"),
        total=Decimal("1210"),
    )
    db.session.add(f)
    db.session.flush()
    db.session.add(
        FacturaItem(
            factura_id=f.id,
            articulo_id=articulo_uno.id,
            codigo=articulo_uno.codigo,
            descripcion=articulo_uno.descripcion,
            cantidad=Decimal("1"),
            precio_unitario=Decimal("1000"),
            descuento_porc=Decimal("0"),
            iva_porc=Decimal("21"),
            iva_monto=Decimal("210"),
            subtotal=Decimal("1000"),
            total=Decimal("1210"),
            orden=0,
        )
    )
    db.session.add(
        FacturaPago(factura_id=f.id, medio=MedioPagoEnum.efectivo, monto=Decimal("1210"), orden=0)
    )
    db.session.commit()
    return f


@pytest.fixture
def compra_uno(db, admin_user, sucursal_uno, articulo_uno, proveedor_uno):
    fecha = datetime.now(timezone.utc).replace(microsecond=0) - timedelta(days=1)
    f = Factura(
        sucursal_id=sucursal_uno.id,
        punto_venta=1,
        tipo=TipoComprobanteEnum.factura_c,
        numero=1,
        fecha=fecha,
        cliente_id=None,
        cajero_id=admin_user.id,
        estado=EstadoComprobanteEnum.emitida,
        subtotal=Decimal("500"),
        total_iva=Decimal("105"),
        total=Decimal("605"),
        legacy_meta={"origen": "ocr", "proveedor_id": proveedor_uno.id, "numero_proveedor": "0001-00000123"},
    )
    db.session.add(f)
    db.session.flush()
    db.session.add(
        FacturaItem(
            factura_id=f.id,
            articulo_id=articulo_uno.id,
            codigo=articulo_uno.codigo,
            descripcion=articulo_uno.descripcion,
            cantidad=Decimal("5"),
            precio_unitario=Decimal("100"),
            descuento_porc=Decimal("0"),
            iva_porc=Decimal("21"),
            iva_monto=Decimal("105"),
            subtotal=Decimal("500"),
            total=Decimal("605"),
            orden=0,
        )
    )
    db.session.commit()
    return f


@pytest.fixture
def cobranza_uno(db, admin_user, sucursal_uno, cliente_uno, venta_uno):
    fecha = datetime.now(timezone.utc).replace(microsecond=0)
    m = MovimientoCaja(
        sucursal_id=sucursal_uno.id,
        caja_numero=1,
        fecha_caja=fecha.date(),
        fecha=fecha,
        tipo=TipoMovimientoEnum.cobranza,
        medio=MedioPagoEnum.efectivo,
        monto=Decimal("1210"),
        factura_id=venta_uno.id,
        cliente_id=cliente_uno.id,
        descripcion="Cobranza factura B",
        user_id=admin_user.id,
    )
    db.session.add(m)
    db.session.commit()
    return m


# --- Tests JSON --------------------------------------------------------------


def test_consulta_clientes_paginado(client, admin_token, auth_header, cliente_uno):
    r = client.get("/api/v1/consultas/clientes", headers=auth_header(admin_token))
    assert r.status_code == 200, r.get_json()
    body = r.get_json()
    assert body["entidad"] == "clientes"
    assert body["page"] == 1
    assert body["per_page"] == 50
    assert body["total"] >= 1
    assert any(item["codigo"] == "CC01" for item in body["items"])


def test_consulta_clientes_q_filter(client, admin_token, auth_header, cliente_uno):
    r = client.get(
        "/api/v1/consultas/clientes?q=Consulta", headers=auth_header(admin_token)
    )
    assert r.status_code == 200
    body = r.get_json()
    assert body["total"] >= 1
    assert any("Consulta" in item["razon_social"] for item in body["items"])


def test_consulta_compras_filtra_por_proveedor(
    client, admin_token, auth_header, compra_uno, proveedor_uno
):
    r = client.get(
        f"/api/v1/consultas/compras?proveedor_id={proveedor_uno.id}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200, r.get_json()
    body = r.get_json()
    assert body["total"] == 1
    item = body["items"][0]
    assert item["proveedor"] == "Prov SA"
    assert item["cuit"] == "30222222229"


def test_consulta_ventas_excluye_compras(
    client, admin_token, auth_header, venta_uno, compra_uno
):
    r = client.get("/api/v1/consultas/ventas", headers=auth_header(admin_token))
    assert r.status_code == 200, r.get_json()
    body = r.get_json()
    assert body["total"] == 1  # solo la factura_b, no la factura_c
    assert body["items"][0]["tipo"] == "factura_b"


def test_consulta_cobranzas(client, admin_token, auth_header, cobranza_uno):
    r = client.get("/api/v1/consultas/cobranzas", headers=auth_header(admin_token))
    assert r.status_code == 200, r.get_json()
    body = r.get_json()
    assert body["total"] == 1
    assert body["items"][0]["cliente"] == "Cliente Consulta"


def test_consulta_stock_bajo(client, admin_token, auth_header, articulo_uno):
    # stock=2 < 5 → debería aparecer
    r = client.get(
        "/api/v1/consultas/stock-bajo?minimo=5", headers=auth_header(admin_token)
    )
    assert r.status_code == 200
    body = r.get_json()
    assert body["total"] == 1
    assert body["items"][0]["codigo"] == "ART001"


def test_consulta_entidad_desconocida(client, admin_token, auth_header):
    r = client.get(
        "/api/v1/consultas/inexistente", headers=auth_header(admin_token)
    )
    assert r.status_code == 404


# --- Tests Excel -------------------------------------------------------------


def test_consulta_clientes_xlsx(client, admin_token, auth_header, cliente_uno):
    r = client.get(
        "/api/v1/consultas/clientes.xlsx", headers=auth_header(admin_token)
    )
    assert r.status_code == 200
    assert r.headers["Content-Type"] == XLSX_MIME
    assert r.data[:2] == b"PK"  # ZIP magic
    wb = load_workbook(io.BytesIO(r.data))
    assert "Clientes" in wb.sheetnames


def test_consulta_compras_xlsx(client, admin_token, auth_header, compra_uno):
    r = client.get(
        "/api/v1/consultas/compras.xlsx", headers=auth_header(admin_token)
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.data))
    sheet = wb.active
    headers = [c.value for c in sheet[1]]
    assert "Comprobante" in headers
    assert "Proveedor" in headers


def test_consulta_xlsx_rechaza_cajero(client, cajero_token, auth_header):
    r = client.get(
        "/api/v1/consultas/clientes.xlsx", headers=auth_header(cajero_token)
    )
    assert r.status_code == 403


def test_consulta_requiere_auth(client):
    r = client.get("/api/v1/consultas/clientes")
    assert r.status_code == 401


# --- Reports nuevos ----------------------------------------------------------


def test_compras_export_xlsx_genera(
    client, admin_token, auth_header, compra_uno
):
    desde = (datetime.now(timezone.utc) - timedelta(days=3)).date().isoformat()
    hasta = datetime.now(timezone.utc).date().isoformat()
    r = client.get(
        f"/api/v1/reports/compras-export.xlsx?fecha_desde={desde}&fecha_hasta={hasta}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.data))
    assert "Detalle compras" in wb.sheetnames
    assert "Resumen por proveedor" in wb.sheetnames


def test_cobranzas_export_xlsx_genera(
    client, admin_token, auth_header, cobranza_uno
):
    desde = (datetime.now(timezone.utc) - timedelta(days=3)).date().isoformat()
    hasta = datetime.now(timezone.utc).date().isoformat()
    r = client.get(
        f"/api/v1/reports/cobranzas-export.xlsx?fecha_desde={desde}&fecha_hasta={hasta}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.data))
    assert "Detalle" in wb.sheetnames
    assert "Por medio de pago" in wb.sheetnames


def test_cta_cte_cliente_xlsx_genera(
    client, admin_token, auth_header, cliente_uno, cobranza_uno
):
    r = client.get(
        f"/api/v1/reports/cta-cte-cliente.xlsx?cliente_id={cliente_uno.id}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"


def test_cta_cte_cliente_xlsx_requiere_id(client, admin_token, auth_header):
    r = client.get(
        "/api/v1/reports/cta-cte-cliente.xlsx", headers=auth_header(admin_token)
    )
    assert r.status_code == 422


def test_resumen_clientes_xlsx(client, admin_token, auth_header, cliente_uno):
    r = client.get(
        "/api/v1/reports/resumen-clientes.xlsx", headers=auth_header(admin_token)
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"


def test_stock_valorizado_xlsx(client, admin_token, auth_header, articulo_uno):
    r = client.get(
        "/api/v1/reports/stock-valorizado.xlsx", headers=auth_header(admin_token)
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"


def test_ventas_detallado_xlsx(client, admin_token, auth_header, venta_uno):
    desde = (datetime.now(timezone.utc) - timedelta(days=3)).date().isoformat()
    hasta = datetime.now(timezone.utc).date().isoformat()
    r = client.get(
        f"/api/v1/reports/ventas-detallado.xlsx?fecha_desde={desde}&fecha_hasta={hasta}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.data))
    assert "Detalle" in wb.sheetnames
    assert "Por sucursal" in wb.sheetnames
    assert "Top 50 productos" in wb.sheetnames


def test_compras_export_rechaza_cajero(client, cajero_token, auth_header):
    r = client.get(
        "/api/v1/reports/compras-export.xlsx", headers=auth_header(cajero_token)
    )
    assert r.status_code == 403
