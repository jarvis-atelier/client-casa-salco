"""Tests de los endpoints xlsx (exports para contador)."""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.models.articulo import Articulo, UnidadMedidaEnum
from app.models.cliente import Cliente, CondicionIvaEnum
from app.models.factura import (
    EstadoComprobanteEnum,
    Factura,
    TipoComprobanteEnum,
)
from app.models.factura_item import FacturaItem
from app.models.pago import FacturaPago, MedioPagoEnum
from app.models.stock import StockSucursal
from app.models.sucursal import Sucursal

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --- Fixtures ----------------------------------------------------------------


@pytest.fixture
def contador_user(db):
    from app.models.user import RolEnum, User
    from app.services.auth_service import hash_password

    user = User(
        email="contador@test.example",
        password_hash=hash_password("contador123"),
        nombre="Contador Test",
        rol=RolEnum.contador,
        activo=True,
    )
    db.session.add(user)
    db.session.commit()
    return user


@pytest.fixture
def contador_token(client, contador_user):
    r = client.post(
        "/api/v1/auth/login",
        json={"email": "contador@test.example", "password": "contador123"},
    )
    assert r.status_code == 200, r.get_json()
    return r.get_json()["access_token"]


@pytest.fixture
def sucursal_a(db):
    s = Sucursal(codigo="SA", nombre="Sucursal A", activa=True)
    db.session.add(s)
    db.session.commit()
    return s


@pytest.fixture
def sucursal_b(db):
    s = Sucursal(codigo="SB", nombre="Sucursal B", activa=True)
    db.session.add(s)
    db.session.commit()
    return s


@pytest.fixture
def cliente_ri(db):
    c = Cliente(
        codigo="C001",
        razon_social="Empresa SA",
        cuit="30123456789",
        condicion_iva=CondicionIvaEnum.responsable_inscripto,
    )
    db.session.add(c)
    db.session.commit()
    return c


@pytest.fixture
def articulos_demo(db, sucursal_a, sucursal_b):
    arts = [
        Articulo(
            codigo="P001",
            descripcion="Pan lactal",
            unidad_medida=UnidadMedidaEnum.unidad,
            costo=Decimal("500"),
            pvp_base=Decimal("1000"),
            iva_porc=Decimal("21"),
            activo=True,
            controla_stock=True,
        ),
        Articulo(
            codigo="P002",
            descripcion="Coca cola 2L",
            unidad_medida=UnidadMedidaEnum.unidad,
            costo=Decimal("1500"),
            pvp_base=Decimal("2500"),
            iva_porc=Decimal("21"),
            activo=True,
            controla_stock=True,
        ),
    ]
    db.session.add_all(arts)
    db.session.flush()
    for art in arts:
        for suc in (sucursal_a, sucursal_b):
            db.session.add(
                StockSucursal(
                    articulo_id=art.id,
                    sucursal_id=suc.id,
                    cantidad=Decimal("100"),
                )
            )
    db.session.commit()
    return arts


def _make_factura(
    db,
    *,
    sucursal_id: int,
    cajero_id: int,
    cliente_id: int | None,
    fecha: datetime,
    items: list[tuple[int, str, Decimal, Decimal]],
    pagos: list[tuple[MedioPagoEnum, Decimal]],
    tipo: TipoComprobanteEnum = TipoComprobanteEnum.factura_b,
    numero: int = 1,
) -> Factura:
    subtotal = sum((c * p for _, _, c, p in items), Decimal("0"))
    iva_total = sum((c * p * Decimal("0.21") for _, _, c, p in items), Decimal("0"))
    total = subtotal + iva_total
    factura = Factura(
        sucursal_id=sucursal_id,
        punto_venta=1,
        tipo=tipo,
        numero=numero,
        fecha=fecha,
        cliente_id=cliente_id,
        cajero_id=cajero_id,
        estado=EstadoComprobanteEnum.emitida,
        subtotal=subtotal.quantize(Decimal("0.01")),
        total_iva=iva_total.quantize(Decimal("0.01")),
        total=total.quantize(Decimal("0.01")),
    )
    db.session.add(factura)
    db.session.flush()
    for orden, (art_id, codigo, cantidad, precio) in enumerate(items):
        line_sub = cantidad * precio
        line_iva = line_sub * Decimal("0.21")
        db.session.add(
            FacturaItem(
                factura_id=factura.id,
                articulo_id=art_id,
                codigo=codigo,
                descripcion=f"Articulo {codigo}",
                cantidad=cantidad,
                precio_unitario=precio,
                descuento_porc=Decimal("0"),
                iva_porc=Decimal("21"),
                iva_monto=line_iva.quantize(Decimal("0.0001")),
                subtotal=line_sub.quantize(Decimal("0.0001")),
                total=(line_sub + line_iva).quantize(Decimal("0.0001")),
                orden=orden,
            )
        )
    for orden, (medio, monto) in enumerate(pagos):
        db.session.add(
            FacturaPago(
                factura_id=factura.id,
                medio=medio,
                monto=monto.quantize(Decimal("0.01")),
                orden=orden,
            )
        )
    db.session.commit()
    return factura


@pytest.fixture
def facturas_seed(db, admin_user, sucursal_a, sucursal_b, cliente_ri, articulos_demo):
    """Seed: 4 facturas a lo largo de 3 días en 2 sucursales."""
    art1, art2 = articulos_demo
    now = datetime.now(timezone.utc).replace(microsecond=0)
    base_today = (now - timedelta(days=3)).replace(hour=11, minute=0, second=0)
    yesterday = base_today - timedelta(days=1)
    two_days_ago = base_today - timedelta(days=2)

    f1 = _make_factura(
        db,
        sucursal_id=sucursal_a.id,
        cajero_id=admin_user.id,
        cliente_id=cliente_ri.id,
        fecha=base_today,
        items=[(art1.id, "P001", Decimal("3"), Decimal("1000"))],
        pagos=[(MedioPagoEnum.efectivo, Decimal("3630"))],
        numero=1,
    )
    f2 = _make_factura(
        db,
        sucursal_id=sucursal_a.id,
        cajero_id=admin_user.id,
        cliente_id=None,
        fecha=base_today.replace(hour=18),
        items=[(art2.id, "P002", Decimal("2"), Decimal("2500"))],
        pagos=[(MedioPagoEnum.tarjeta_credito, Decimal("6050"))],
        tipo=TipoComprobanteEnum.ticket,
        numero=2,
    )
    f3 = _make_factura(
        db,
        sucursal_id=sucursal_b.id,
        cajero_id=admin_user.id,
        cliente_id=cliente_ri.id,
        fecha=yesterday,
        items=[
            (art1.id, "P001", Decimal("1"), Decimal("1000")),
            (art2.id, "P002", Decimal("1"), Decimal("2500")),
        ],
        pagos=[(MedioPagoEnum.qr_mercadopago, Decimal("4235"))],
        numero=2,
    )
    f4 = _make_factura(
        db,
        sucursal_id=sucursal_b.id,
        cajero_id=admin_user.id,
        cliente_id=None,
        fecha=two_days_ago,
        items=[(art1.id, "P001", Decimal("5"), Decimal("1000"))],
        pagos=[(MedioPagoEnum.efectivo, Decimal("6050"))],
        tipo=TipoComprobanteEnum.ticket,
        numero=3,
    )
    return [f1, f2, f3, f4]


def _date_range_query(facturas: list[Factura]) -> tuple[str, str]:
    fechas = [f.fecha.date() for f in facturas]
    return min(fechas).isoformat(), max(fechas).isoformat()


# --- Tests --------------------------------------------------------------------


def test_libro_iva_digital_devuelve_xlsx_valido(
    client, admin_token, auth_header, facturas_seed
):
    desde, hasta = _date_range_query(facturas_seed)
    r = client.get(
        f"/api/v1/reports/libro-iva-digital.xlsx?fecha_desde={desde}&fecha_hasta={hasta}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200, r.data[:200]
    assert r.headers["Content-Type"] == XLSX_MIME
    assert "attachment" in r.headers["Content-Disposition"].lower()
    assert ".xlsx" in r.headers["Content-Disposition"]
    # Magic number ZIP
    assert r.data[:2] == b"PK"

    wb = load_workbook(io.BytesIO(r.data))
    assert "Ventas" in wb.sheetnames
    assert "Compras" in wb.sheetnames

    ventas = wb["Ventas"]
    # Cabecera
    headers = [c.value for c in ventas[1]]
    assert "Fecha" in headers
    assert "CUIT" in headers
    assert "Total" in headers
    # Filas: 4 facturas seed
    data_rows = list(ventas.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 4

    # Compras: solo cabecera (vacío hoy)
    compras = wb["Compras"]
    compras_rows = list(compras.iter_rows(min_row=2, values_only=True))
    assert compras_rows == []  # vacío


def test_ventas_export_filtra_por_sucursal(
    client, admin_token, auth_header, sucursal_a, facturas_seed
):
    desde, hasta = _date_range_query(facturas_seed)
    r = client.get(
        f"/api/v1/reports/ventas-export.xlsx?fecha_desde={desde}&fecha_hasta={hasta}"
        f"&sucursal_id={sucursal_a.id}",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200, r.data[:200]

    wb = load_workbook(io.BytesIO(r.data))
    assert "Detalle" in wb.sheetnames
    assert "Resumen Diario" in wb.sheetnames

    detalle = wb["Detalle"]
    rows = list(detalle.iter_rows(min_row=2, values_only=True))
    # Sucursal A tiene f1 + f2 = 2 facturas
    assert len(rows) == 2
    # Columna 2 (Sucursal): debe contener "SA" en todas
    for row in rows:
        assert "SA" in (row[1] or "")


def test_stock_export_genera_filas_para_todos_los_articulos(
    client, admin_token, auth_header, articulos_demo, sucursal_a, sucursal_b
):
    r = client.get(
        "/api/v1/reports/stock-export.xlsx",
        headers=auth_header(admin_token),
    )
    assert r.status_code == 200, r.data[:200]

    wb = load_workbook(io.BytesIO(r.data))
    assert "Stock" in wb.sheetnames
    ws = wb["Stock"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # 2 artículos × 2 sucursales = 4 filas
    assert len(rows) == 4
    # Columnas: codigo, descripcion, unidad, sucursal, cantidad, costo, valor, pvp
    codigos = {r[0] for r in rows}
    assert codigos == {"P001", "P002"}


def test_libro_iva_xlsx_rechaza_cajero(
    client, cajero_token, auth_header
):
    r = client.get(
        "/api/v1/reports/libro-iva-digital.xlsx",
        headers=auth_header(cajero_token),
    )
    assert r.status_code == 403, r.data[:200]


def test_contador_puede_exportar_libro_iva(
    client, contador_token, auth_header, facturas_seed
):
    desde, hasta = _date_range_query(facturas_seed)
    r = client.get(
        f"/api/v1/reports/libro-iva-digital.xlsx?fecha_desde={desde}&fecha_hasta={hasta}",
        headers=auth_header(contador_token),
    )
    assert r.status_code == 200
    assert r.data[:2] == b"PK"


def test_exports_requieren_auth(client):
    r = client.get("/api/v1/reports/libro-iva-digital.xlsx")
    assert r.status_code == 401
    r = client.get("/api/v1/reports/ventas-export.xlsx")
    assert r.status_code == 401
    r = client.get("/api/v1/reports/stock-export.xlsx")
    assert r.status_code == 401
